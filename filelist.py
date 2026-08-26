#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
目录清单生成器（独立新实现）
============================
一个放进任意文件夹就能用的 Excel 文件清单工具：
运行后自动扫描【脚本所在文件夹】，生成 文件清单.xlsx。

[双击运行]
    把本脚本和 生成文件清单.bat 放在一起，双击 bat。
    依次回答 3 个问题（直接回车 = 用默认值）：
      1. 文件名要可点击的超链接吗？（默认：是）
      2. 要列出隐藏文件和隐藏文件夹吗？（默认：否）
      3. 要递归扫描子文件夹吗？（默认：是）

[命令行运行]
    python filelist.py [选项]

选项：
    -s, --source     要扫描的文件夹（默认：脚本所在文件夹）
    -o, --output     输出文件路径（默认：被扫描文件夹/文件清单.xlsx）
        --links      文件名生成可点击超链接（相对路径，默认开启）
        --no-links   不生成超链接
        --hidden     列出隐藏文件和隐藏文件夹
        --no-hidden  不列出隐藏文件和隐藏文件夹（默认关闭）
        --recursive  递归进入子文件夹（默认开启）
        --no-recursive 只扫描顶层

示例：
    python filelist.py
    python filelist.py --no-hidden
    python filelist.py --no-links --no-hidden
    python filelist.py -s "D:/下载" -o "D:/下载/我的清单.xlsx"

说明：
    - 自动跳过工具自身文件（filelist.py、生成文件清单.bat）和本次输出文件。
    - 输出文件被 Excel 占用时，自动另存为带时间戳的文件。
    - 其他电脑可用：需要 Python 3；缺 openpyxl 时会自动安装（需联网）。
    - 超链接只加在“文件名”列，目标用相对路径（以清单文件所在文件夹为基准）。
    - 打包成 exe 后用法相同：放到任意文件夹双击，即扫描 exe 所在文件夹。
"""

import argparse
import datetime
import os
import stat
import sys
from collections import namedtuple
from urllib.parse import quote

# ---------------------------------------------------------------- 常量
MAX_SHEET_ROWS = 1_048_575        # Excel 单工作表行数上限
LINK_SAFE_LIMIT = 32_000          # 超过该文件数时自动关闭超链接（防文件损坏）
MAX_LINK_LENGTH = 255             # Excel 超链接 Target 长度上限
LONG_PATH_HINT = "[路径过长，请手动打开]"

# 一条文件记录
FileInfo = namedtuple("FileInfo", "name ext size size_text mtime folder path")


def console_safe():
    """不同电脑的控制台编码不同，这里兜底，避免打印中文时崩溃。"""
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if reconfigure is not None:
            try:
                reconfigure(errors="replace")
            except Exception:
                pass


def ensure_openpyxl(interactive):
    """拿到 openpyxl；没有就先问/装一个。"""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        if interactive:
            answer = input("缺少 openpyxl，是否自动安装？(Y/n，回车=是)：").strip().lower()
            if answer in ("n", "no", "否"):
                print("已取消。请手动执行：pip install openpyxl")
                sys.exit(1)
        print("正在安装 openpyxl ……")
        import subprocess
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", "--user", "openpyxl"]
            )
            import openpyxl
            return openpyxl
        except Exception as exc:
            print(f"自动安装失败：{exc}")
            print("请手动执行：pip install --user openpyxl")
            sys.exit(1)


def format_size(num_bytes):
    """把字节数换算成人能看懂的大小。"""
    if num_bytes < 1024:
        return f"{num_bytes} B"
    value = float(num_bytes)
    for unit in ("KB", "MB", "GB", "TB"):
        value /= 1024.0
        if value < 1024:
            return f"{value:.2f} {unit}"
    return f"{value:.2f} PB"


def is_hidden(path):
    """判断是否为隐藏：点开头，或 Windows 带隐藏属性。"""
    name = os.path.basename(path.rstrip("\\/"))
    if name.startswith("."):
        return True
    if os.name == "nt":
        try:
            attributes = os.stat(path).st_file_attributes
            return bool(attributes & getattr(stat, "FILE_ATTRIBUTE_HIDDEN", 0x2))
        except (AttributeError, OSError):
            pass
    return False


def hyperlink_target(path):
    """把本地路径转成 Excel 可点击的 file:/// 链接。"""
    normalized = os.path.abspath(path).replace("\\", "/")
    uri = "file:///" + quote(normalized, safe=":/")
    truncated = len(uri) > MAX_LINK_LENGTH
    return (uri[:MAX_LINK_LENGTH] if truncated else uri), truncated


def relative_target(full, base_dir):
    """生成相对路径超链接：以清单文件所在目录为基准。

    中文等字符无需百分号编码，链接短、可点击；跨盘符时返回 None。
    仅编码会影响 URI 的 ASCII 字符（空格、#、%、?、&、引号等）。
    """
    try:
        rel = os.path.relpath(full, base_dir)
    except ValueError:
        return None
    rel = rel.replace("\\", "/")
    encoded = []
    for ch in rel:
        if ch in " %#?&\"'<>|":
            encoded.append(quote(ch, safe=""))
        else:
            encoded.append(ch)
    return "".join(encoded)


def _make_file_info(root, full, name):
    """取单个文件的信息；失败返回 None。"""
    try:
        st = os.stat(full)
    except OSError:
        return None
    folder = os.path.relpath(os.path.dirname(full), root)
    if folder == ".":
        folder = "(根目录)"
    ext = os.path.splitext(name)[1].lower() or "(无扩展名)"
    return FileInfo(
        name=name,
        ext=ext,
        size=st.st_size,
        size_text=format_size(st.st_size),
        mtime=datetime.datetime.fromtimestamp(st.st_mtime),
        folder=folder,
        path=full,
    )


def scan_folder(root, include_hidden, recursive, exclude, errors, skip_names=None):
    """扫描文件夹。

    返回 (文件列表, 文件夹数量)。
    用显式栈代替系统递归，深层目录也不会超出 Python 递归上限；
    符号链接/联接只记录不进入，防止绕出根目录或死循环。
    """
    found = []
    folder_count = 0
    stack = [root]
    while stack:
        current = stack.pop()
        folder_count += 1
        try:
            entries = list(os.scandir(current))
        except OSError as exc:
            errors.append(f"无法访问：{current}（{exc}）")
            continue
        for entry in entries:
            full = entry.path
            if os.path.normcase(os.path.abspath(full)) in exclude:
                continue
            if skip_names and entry.name.lower() in skip_names:
                continue
            if not include_hidden and is_hidden(full):
                continue
            if entry.is_symlink():
                info = _make_file_info(root, full, entry.name)
                if info is not None:
                    found.append(info)
                continue
            try:
                if entry.is_dir(follow_symlinks=False):
                    if recursive:
                        stack.append(full)
                elif entry.is_file(follow_symlinks=False):
                    info = _make_file_info(root, full, entry.name)
                    if info is not None:
                        found.append(info)
            except OSError:
                continue
    found.sort(key=lambda item: (item.folder, item.name.lower()))
    return found, folder_count


def write_excel(openpyxl, files, folder_count, options, errors, link_base=None):
    """把扫描结果写成 Excel 工作簿。"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "文件清单"

    Font = openpyxl.styles.Font
    PatternFill = openpyxl.styles.PatternFill
    Alignment = openpyxl.styles.Alignment
    Border = openpyxl.styles.Border
    Side = openpyxl.styles.Side
    get_column_letter = openpyxl.utils.get_column_letter

    headers = ["序号", "文件名", "扩展名", "文件大小", "修改日期", "所在文件夹", "完整路径"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_side = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    link_font = Font(color="0563C1", underline="single")

    for column in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    use_links = options["links"] and len(files) <= LINK_SAFE_LIMIT
    if options["links"] and not use_links:
        print(f"[提示] 文件数（{len(files)}）过多，已自动关闭超链接以保证文件安全。")

    for index, item in enumerate(files, start=1):
        sheet.append([
            index, item.name, item.ext, item.size_text, item.mtime,
            item.folder, item.path,
        ])
        row_index = index + 1
        for column in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_index, column=column)
            cell.border = border
            cell.alignment = center if column in (1, 3) else left
            if column == 5:
                cell.number_format = "YYYY-MM-DD HH:MM"
        if use_links:
            target = None
            if link_base is not None:
                candidate = relative_target(item.path, link_base)
                if candidate is not None and len(candidate) <= MAX_LINK_LENGTH:
                    target = candidate
            truncated = False
            if target is None:
                target, truncated = hyperlink_target(item.path)
            name_cell = sheet.cell(row=row_index, column=2)
            path_cell = sheet.cell(row=row_index, column=7)
            if truncated:
                path_cell.value = f"{item.path}\n{LONG_PATH_HINT}"
                path_cell.alignment = left
            else:
                # 只在“文件名”列加超链接，目标用相对路径（以清单所在文件夹为基准）
                name_cell.hyperlink = target
                name_cell.font = link_font
        if index % 2 == 0:
            stripe = PatternFill("solid", fgColor="F2F6FC")
            for column in range(1, len(headers) + 1):
                sheet.cell(row=row_index, column=column).fill = stripe

    widths = [6, 48, 12, 14, 18, 30, 60]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(files) + 1}"

    # ---- 统计汇总 ----
    summary = wb.create_sheet("统计汇总")
    summary.append(["项目", "数值"])
    for column in (1, 2):
        cell = summary.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    total_size = sum(item.size for item in files)
    stats = [
        ("文件总数", len(files)),
        ("文件夹总数", folder_count),
        ("总大小", format_size(total_size)),
        ("扫描时间", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("超链接", "是" if use_links else "否"),
        ("隐藏文件", "是" if options["hidden"] else "否"),
        ("递归扫描", "是" if options["recursive"] else "否"),
    ]
    for key, value in stats:
        summary.append([key, value])
        row = summary.max_row
        summary.cell(row=row, column=1).border = border
        summary.cell(row=row, column=2).border = border

    summary.append([])
    summary.append(["扩展名", "文件数量"])
    header_row = summary.max_row
    green_fill = PatternFill("solid", fgColor="70AD47")
    for column in (1, 2):
        cell = summary.cell(row=header_row, column=column)
        cell.fill = green_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = center
        cell.border = border

    ext_count = {}
    for item in files:
        ext_count[item.ext] = ext_count.get(item.ext, 0) + 1
    for ext, count in sorted(ext_count.items(), key=lambda pair: -pair[1]):
        summary.append([ext, count])
        row = summary.max_row
        summary.cell(row=row, column=1).border = border
        summary.cell(row=row, column=2).border = border

    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 24

    if errors:
        print(f"[注意] 有 {len(errors)} 个目录无法访问（已跳过）：")
        for message in errors[:10]:
            print(f"       {message}")
        if len(errors) > 10:
            print(f"       …… 共 {len(errors)} 条")

    return wb


def save_workbook(wb, target):
    """保存；目标被占用时自动另存为带时间戳的文件。"""
    try:
        wb.save(target)
        return target
    except PermissionError:
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        directory = os.path.dirname(os.path.abspath(target))
        fallback = os.path.join(directory, f"文件清单_{stamp}.xlsx")
        print(f"[提示] {target} 正被占用（可能已在 Excel 中打开），另存为：{fallback}")
        wb.save(fallback)
        return fallback


def main():
    console_safe()

    parser = argparse.ArgumentParser(
        description="生成文件夹文件清单 (Excel)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("-s", "--source", default=None,
                        help="要扫描的文件夹（默认：脚本所在文件夹）")
    parser.add_argument("-o", "--output", default=None,
                        help="输出文件路径（默认：被扫描文件夹/文件清单.xlsx）")

    group_links = parser.add_mutually_exclusive_group()
    group_links.add_argument("--links", dest="links", action="store_true",
                             help="文件名生成可点击超链接（相对路径，默认）")
    group_links.add_argument("--no-links", dest="links", action="store_false",
                             help="不生成超链接")
    parser.set_defaults(links=None)

    group_hidden = parser.add_mutually_exclusive_group()
    group_hidden.add_argument("--hidden", dest="hidden", action="store_true",
                              help="列出隐藏文件和隐藏文件夹")
    group_hidden.add_argument("--no-hidden", dest="hidden", action="store_false",
                              help="不列出隐藏文件和隐藏文件夹（默认）")
    parser.set_defaults(hidden=None)

    group_recursive = parser.add_mutually_exclusive_group()
    group_recursive.add_argument("--recursive", dest="recursive", action="store_true",
                                 help="递归扫描子文件夹（默认）")
    group_recursive.add_argument("--no-recursive", dest="recursive", action="store_false",
                                 help="只扫描顶层")
    parser.set_defaults(recursive=None)

    args = parser.parse_args()

    interactive = sys.stdin.isatty()

    def ask(question, default=True):
        """双击运行时提问；命令行/管道直接返回默认值。"""
        if not interactive:
            return default
        try:
            answer = input(question).strip().lower()
        except EOFError:
            return default
        if answer == "":
            return default
        return answer not in ("n", "no", "否", "f")

    # 打包成 exe 后，__file__ 指向临时解压目录，须改用 exe 所在文件夹
    if getattr(sys, "frozen", False):
        script_dir = os.path.dirname(os.path.abspath(sys.executable))
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
    source = os.path.abspath(args.source or script_dir)
    links = args.links if args.links is not None else ask("文件名要可点击的超链接吗？(Y/否，回车=是)：")
    hidden = args.hidden if args.hidden is not None else ask("要列出隐藏文件和隐藏文件夹吗？(Y/是，回车=否)：", default=False)
    recursive = args.recursive if args.recursive is not None else ask("要递归扫描子文件夹吗？(Y/否，回车=是)：")
    output = os.path.abspath(args.output or os.path.join(source, "文件清单.xlsx"))

    if not os.path.isdir(source):
        print(f"[错误] 文件夹不存在：{source}")
        sys.exit(1)

    options = {"links": links, "hidden": hidden, "recursive": recursive}

    print(f"正在扫描：{source}")
    print(f"  超链接 = {'是' if links else '否'}，"
          f"隐藏文件 = {'是' if hidden else '否'}，"
          f"递归 = {'是' if recursive else '否'}")

    openpyxl = ensure_openpyxl(interactive)

    # 工具自身（脚本/exe + 启动器 bat）不列入清单
    own_path = sys.executable if getattr(sys, "frozen", False) else __file__

    # 排除：工具自身 + 本次输出文件（大小写不敏感的绝对路径）
    exclude = {
        os.path.normcase(os.path.abspath(own_path)),
        os.path.normcase(os.path.abspath(output)),
    }
    skip_names = {
        os.path.basename(os.path.abspath(own_path)).lower(),
        "生成文件清单.bat",
    }

    errors = []
    files, folder_count = scan_folder(source, hidden, recursive, exclude, errors, skip_names)

    # 超过 Excel 单工作表行数上限时截断，避免保存失败
    if len(files) > MAX_SHEET_ROWS:
        print(f"[警告] 文件数（{len(files)}）超过 Excel 单工作表上限（{MAX_SHEET_ROWS}），已截断至前 {MAX_SHEET_ROWS} 条。")
        files = files[:MAX_SHEET_ROWS]
    wb = write_excel(openpyxl, files, folder_count, options, errors, link_base=os.path.dirname(output))
    saved = save_workbook(wb, output)

    print("完成！")
    print(f"  文件总数：{len(files)}")
    print(f"  文件夹总数：{folder_count}")
    print(f"  总大小：{format_size(sum(item.size for item in files))}")
    print(f"  已保存：{saved}")


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        raise
    except Exception as exc:
        import traceback
        print(f"\n[错误] 运行出错：{exc}")
        traceback.print_exc()
    finally:
        if sys.stdin.isatty():
            try:
                input("\n按回车键退出...")
            except (EOFError, KeyboardInterrupt):
                pass

